import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import logging
from database import db

logger = logging.getLogger(__name__)

class ExcelExportService:
    def __init__(self):
        self.export_dir = "user_exports"
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_users_to_excel(self, filename=None):
        """Export all users to Excel file"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"wealthsage_users_{timestamp}.xlsx"
            
            filepath = os.path.join(self.export_dir, filename)
            
            # Get all users from database
            users_data = self.get_all_users_data()
            
            if not users_data:
                logger.warning("No users found to export")
                return None
            
            # Create DataFrame
            df = pd.DataFrame(users_data)
            
            # Create Excel workbook with multiple sheets
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Main users sheet
                df.to_excel(writer, sheet_name='Users', index=False)
                
                # Get workbook and worksheet for styling
                workbook = writer.book
                worksheet = workbook['Users']
                
                # Style the header
                self.style_header(worksheet)
                
                # Auto-adjust column widths
                self.adjust_column_widths(worksheet)
                
                # Add summary sheet
                self.create_summary_sheet(workbook, users_data)
                
                # Add statistics sheet
                self.create_statistics_sheet(workbook, users_data)
            
            logger.info(f"Users exported successfully to {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting users to Excel: {e}")
            return None
    
    def get_all_users_data(self):
        """Get all users data from database"""
        try:
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT 
                        u.id,
                        u.email,
                        u.first_name,
                        u.last_name,
                        u.display_name,
                        u.role,
                        u.phone,
                        u.is_active,
                        u.is_verified,
                        u.created_at,
                        u.last_login,
                        u.provider,
                        p.university,
                        p.monthly_income,
                        p.savings_goal,
                        p.current_savings,
                        p.budget_limit
                    FROM users u
                    LEFT JOIN user_profiles p ON u.id = p.user_id
                    ORDER BY u.created_at DESC
                ''')
                
                columns = [
                    'ID', 'Email', 'First Name', 'Last Name', 'Display Name', 
                    'Role', 'Phone', 'Active', 'Verified', 'Created At', 
                    'Last Login', 'Provider', 'University', 'Monthly Income',
                    'Savings Goal', 'Current Savings', 'Budget Limit'
                ]
                
                rows = cursor.fetchall()
                users_data = []
                
                for row in rows:
                    user_dict = {}
                    for i, column in enumerate(columns):
                        value = row[i] if i < len(row) else None
                        
                        # Format specific fields
                        if column in ['Active', 'Verified'] and value is not None:
                            user_dict[column] = 'Yes' if value else 'No'
                        elif column in ['Monthly Income', 'Savings Goal', 'Current Savings', 'Budget Limit']:
                            user_dict[column] = f"${value:.2f}" if value else "$0.00"
                        elif column in ['Created At', 'Last Login'] and value:
                            # Format datetime
                            if isinstance(value, str):
                                try:
                                    dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                                    user_dict[column] = dt.strftime("%Y-%m-%d %H:%M:%S")
                                except:
                                    user_dict[column] = value
                            else:
                                user_dict[column] = str(value)
                        else:
                            user_dict[column] = value or ''
                    
                    users_data.append(user_dict)
                
                return users_data
                
        except Exception as e:
            logger.error(f"Error getting users data: {e}")
            return []
    
    def style_header(self, worksheet):
        """Style the header row"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def adjust_column_widths(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_sheet(self, workbook, users_data):
        """Create summary sheet with key metrics"""
        summary_sheet = workbook.create_sheet("Summary")
        
        # Calculate summary statistics
        total_users = len(users_data)
        active_users = sum(1 for user in users_data if user.get('Active') == 'Yes')
        verified_users = sum(1 for user in users_data if user.get('Verified') == 'Yes')
        
        # Role distribution
        role_counts = {}
        for user in users_data:
            role = user.get('Role', 'Unknown')
            role_counts[role] = role_counts.get(role, 0) + 1
        
        # Provider distribution
        provider_counts = {}
        for user in users_data:
            provider = user.get('Provider', 'Unknown')
            provider_counts[provider] = provider_counts.get(provider, 0) + 1
        
        # Write summary data
        summary_data = [
            ['WealthSage User Summary', ''],
            ['Generated On', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ['', ''],
            ['Total Users', total_users],
            ['Active Users', active_users],
            ['Verified Users', verified_users],
            ['', ''],
            ['Role Distribution', ''],
        ]
        
        for role, count in role_counts.items():
            summary_data.append([f'  {role}', count])
        
        summary_data.extend([
            ['', ''],
            ['Provider Distribution', ''],
        ])
        
        for provider, count in provider_counts.items():
            summary_data.append([f'  {provider}', count])
        
        # Write to sheet
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Style headers
                if row_idx == 1:
                    cell.font = Font(bold=True, size=16)
                elif value in ['Role Distribution', 'Provider Distribution', 'WealthSage User Summary']:
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        self.adjust_column_widths(summary_sheet)
    
    def create_statistics_sheet(self, workbook, users_data):
        """Create statistics sheet with detailed analytics"""
        stats_sheet = workbook.create_sheet("Statistics")
        
        # Registration trends (by month)
        monthly_registrations = {}
        for user in users_data:
            created_at = user.get('Created At', '')
            if created_at:
                try:
                    month = created_at[:7]  # YYYY-MM format
                    monthly_registrations[month] = monthly_registrations.get(month, 0) + 1
                except:
                    pass
        
        # Financial data analysis
        total_savings_goals = 0
        total_current_savings = 0
        users_with_goals = 0
        
        for user in users_data:
            savings_goal = user.get('Savings Goal', '$0.00')
            current_savings = user.get('Current Savings', '$0.00')
            
            try:
                goal_amount = float(savings_goal.replace('$', '').replace(',', ''))
                current_amount = float(current_savings.replace('$', '').replace(',', ''))
                
                if goal_amount > 0:
                    total_savings_goals += goal_amount
                    users_with_goals += 1
                
                total_current_savings += current_amount
            except:
                pass
        
        # Write statistics
        stats_data = [
            ['WealthSage Detailed Statistics', ''],
            ['', ''],
            ['Registration Trends', ''],
        ]
        
        for month, count in sorted(monthly_registrations.items()):
            stats_data.append([f'  {month}', count])
        
        stats_data.extend([
            ['', ''],
            ['Financial Overview', ''],
            ['Total Savings Goals', f'${total_savings_goals:,.2f}'],
            ['Total Current Savings', f'${total_current_savings:,.2f}'],
            ['Users with Savings Goals', users_with_goals],
            ['Average Savings Goal', f'${total_savings_goals/max(users_with_goals, 1):,.2f}'],
            ['Average Current Savings', f'${total_current_savings/max(len(users_data), 1):,.2f}'],
        ])
        
        # Write to sheet
        for row_idx, row_data in enumerate(stats_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = stats_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Style headers
                if row_idx == 1:
                    cell.font = Font(bold=True, size=16)
                elif value in ['Registration Trends', 'Financial Overview']:
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        self.adjust_column_widths(stats_sheet)
    
    def export_user_on_signup(self, user_data):
        """Export single user data when they sign up"""
        try:
            # Create or update the main users file
            main_file = os.path.join(self.export_dir, "wealthsage_all_users.xlsx")
            
            # Prepare user data for Excel
            excel_data = {
                'ID': user_data.get('id', ''),
                'Email': user_data.get('email', ''),
                'First Name': user_data.get('first_name', ''),
                'Last Name': user_data.get('last_name', ''),
                'Display Name': user_data.get('display_name', ''),
                'Role': user_data.get('role', ''),
                'Phone': user_data.get('phone', ''),
                'Active': 'Yes' if user_data.get('is_active') else 'No',
                'Verified': 'Yes' if user_data.get('is_verified') else 'No',
                'Created At': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'Provider': user_data.get('provider', 'email'),
                'University': user_data.get('university', ''),
                'Monthly Income': '$0.00',
                'Savings Goal': '$0.00',
                'Current Savings': '$0.00',
                'Budget Limit': '$0.00'
            }
            
            # Check if file exists
            if os.path.exists(main_file):
                # Append to existing file
                df_existing = pd.read_excel(main_file)
                df_new = pd.DataFrame([excel_data])
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            else:
                # Create new file
                df_combined = pd.DataFrame([excel_data])
            
            # Save to Excel
            df_combined.to_excel(main_file, index=False)
            
            logger.info(f"User {user_data.get('email')} exported to Excel")
            return main_file
            
        except Exception as e:
            logger.error(f"Error exporting user on signup: {e}")
            return None

# Global instance
excel_service = ExcelExportService()
