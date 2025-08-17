from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import firebase_admin
from firebase_admin import credentials, auth
import os
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import logging

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

# Initialize Firebase Admin with environment variable
try:
    # Get Firebase service account path from environment
    firebase_service_account_path = os.getenv('FIREBASE_SERVICE_ACCOUNT_PATH')
    
    if not firebase_service_account_path:
        logger.error("FIREBASE_SERVICE_ACCOUNT_PATH not found in environment variables")
        raise ValueError("Firebase service account path not configured")
    
    if not os.path.exists(firebase_service_account_path):
        logger.error(f"Firebase service account file not found: {firebase_service_account_path}")
        raise FileNotFoundError(f"Firebase service account file not found: {firebase_service_account_path}")
    
    cred = credentials.Certificate(firebase_service_account_path)
    firebase_admin.initialize_app(cred)
    logger.info("Firebase Admin initialized successfully")
except Exception as e:
    logger.error(f"Failed to initialize Firebase Admin: {e}")
    raise

# Configure SQLAlchemy - Choose your database
# For MySQL: mysql://username:password@localhost/database_name
# For PostgreSQL: postgresql://username:password@localhost/database_name
# For SQLite (development): sqlite:///users.db

DATABASE_URL = os.getenv('DATABASE_URL', 'sqlite:///users.db')
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# User Model with role and timestamps
class User(db.Model):
    __tablename__ = 'users'
    
    # Primary key - Firebase UID
    uid = db.Column(db.String(128), primary_key=True)
    
    # User details
    email = db.Column(db.String(120), nullable=False)  # Removed unique constraint
    name = db.Column(db.String(120), nullable=False)
    picture = db.Column(db.String(255))
    role = db.Column(db.String(50), default='user')  # Default role
    
    # Timestamps
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        """Convert user object to dictionary"""
        return {
            'uid': self.uid,
            'email': self.email,
            'name': self.name,
            'picture': self.picture,
            'role': self.role,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'updated_at': self.updated_at.isoformat() if self.updated_at else None
        }

@app.route('/api/auth/google', methods=['POST'])
def google_auth():
    """
    Handle Google Auth signup/login
    - Verify Firebase ID token
    - Store user in SQL database if new
    - Return user data
    """
    try:
        # Get the ID token from the request
        data = request.get_json()
        id_token = data.get('idToken')
        role = data.get('role', 'user')  # Get role from request, default to 'user'
        
        if not id_token:
            return jsonify({'error': 'No ID token provided'}), 400

        logger.info(f"Processing Google Auth with role: {role}")

        # Verify the ID token with Firebase
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token['uid']
        email = decoded_token.get('email')
        name = decoded_token.get('name')
        picture = decoded_token.get('picture')

        logger.info(f"Token verified for user: {email}")

        # Check if user already exists in SQL database by UID
        existing_user = db.session.get(User, uid)
        
        if existing_user:
            logger.info(f"User {email} already exists, updating last login")
            # Update last login time
            existing_user.updated_at = datetime.utcnow()
            db.session.commit()
            
            return jsonify({
                'success': True,
                'user': existing_user.to_dict(),
                'isNewUser': False
            })
        else:
            # Check if user exists with same email but different UID (email/password user)
            existing_email_user = db.session.query(User).filter(User.email == email).first()
            
            if existing_email_user:
                logger.info(f"User with email {email} exists but different UID. Updating UID and profile.")
                # Update the existing user's UID to the Google UID and update profile
                existing_email_user.uid = uid
                existing_email_user.name = name
                existing_email_user.picture = picture
                existing_email_user.updated_at = datetime.utcnow()
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'user': existing_email_user.to_dict(),
                    'isNewUser': False
                })
            else:
                # Create new user in SQL database
                logger.info(f"Creating new user: {email}")
                new_user = User(
                    uid=uid,
                    email=email,
                    name=name,
                    picture=picture,
                    role=role,
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow()
                )
                
                db.session.add(new_user)
                db.session.commit()
                
                logger.info(f"New user created successfully: {email}")
                
                return jsonify({
                    'success': True,
                    'user': new_user.to_dict(),
                    'isNewUser': True
                })

    except Exception as e:
        logger.error(f"Error in Google Auth: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users', methods=['GET'])
def get_users():
    """
    Get all users (admin only)
    """
    try:
        users = db.session.query(User).all()
        return jsonify({
            'success': True,
            'users': [user.to_dict() for user in users],
            'count': len(users)
        })
    except Exception as e:
        logger.error(f"Error fetching users: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-users', methods=['GET'])
def export_users():
    """
    Export all users to Excel file
    - Fetch all users from SQL database
    - Generate Excel file with user details
    - Return file as download
    """
    try:
        logger.info("Starting user export to Excel")
        
        # Fetch all users from database
        users = db.session.query(User).all()
        logger.info(f"Found {len(users)} users to export")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Add headers
        headers = ['UID', 'Name', 'Email', 'Role', 'Created At', 'Updated At']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add user data
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.uid)
            ws.cell(row=row, column=2, value=user.name)
            ws.cell(row=row, column=3, value=user.email)
            ws.cell(row=row, column=4, value=user.role)
            ws.cell(row=row, column=5, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '')
            ws.cell(row=row, column=6, value=user.updated_at.strftime('%Y-%m-%d %H:%M:%S') if user.updated_at else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        logger.info("Excel file generated successfully")
        
        # Return file as download
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'users_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Error exporting users: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<uid>', methods=['PUT'])
def update_user_role(uid):
    """
    Update user role (admin only)
    """
    try:
        data = request.get_json()
        new_role = data.get('role')
        
        if not new_role:
            return jsonify({'error': 'Role is required'}), 400
        
        user = db.session.get(User, uid)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        user.role = new_role
        user.updated_at = datetime.utcnow()
        db.session.commit()
        
        logger.info(f"Updated role for user {user.email} to {new_role}")
        
        return jsonify({
            'success': True,
            'user': user.to_dict()
        })
        
    except Exception as e:
        logger.error(f"Error updating user role: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat(),
        'database': 'connected' if db.engine.pool.checkedin() >= 0 else 'disconnected'
    })

# Create database tables
def init_db():
    """Initialize database tables"""
    try:
        with app.app_context():
            db.create_all()
            logger.info("Database tables created successfully")
    except Exception as e:
        logger.error(f"Error creating database tables: {e}")

if __name__ == '__main__':
    # Initialize database
    init_db()
    
    # Run the app
    port = int(os.getenv('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
